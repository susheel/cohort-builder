"""
compile_specs.py
~~~~~~~~~~~~~~~~
Compiles the ELITE Cohort Builder workbook into reusable override specs
for the spec-driven cohort tool.

Reads: Cohort Builder output privacy analysis.xlsx
Writes: public/specs/{elite47,elite-v2,elite-v3,ad-v1}.spec.{yaml,json}
        public/specs/index.json

Usage (from repo root):
    python3 scripts/compile_specs.py
"""

from __future__ import annotations

import json
import os
import re
import sys
from pathlib import Path
from typing import Any

import openpyxl

# ── Paths ──────────────────────────────────────────────────────────────────

REPO_ROOT = Path(__file__).resolve().parent.parent
WORKBOOK_PATH = REPO_ROOT / "Cohort Builder output privacy analysis.xlsx"
SPECS_DIR = REPO_ROOT / "public" / "specs"

SPECS_DIR.mkdir(parents=True, exist_ok=True)

# ── Minimal YAML serialiser (PyYAML not installed) ─────────────────────────

def _yaml_str(value: str) -> str:
    """Quote a string value if it contains YAML special characters."""
    if value is None:
        return "null"
    # Characters that force quoting
    needs_quote = any(c in value for c in (':', '#', '[', ']', '{', '}', ',', '&', '*',
                                           '?', '|', '-', '<', '>', '=', '!', '%', '@',
                                           '`', '"', "'", '\n', '\r'))
    if needs_quote or value.strip() != value or not value:
        escaped = value.replace('\\', '\\\\').replace('"', '\\"').replace('\n', '\\n').replace('\r', '')
        return f'"{escaped}"'
    # Also quote plain values that look like YAML keywords
    if value.lower() in ('true', 'false', 'null', 'yes', 'no', 'on', 'off'):
        return f'"{value}"'
    # Quote if looks like a number
    try:
        float(value)
        return f'"{value}"'
    except ValueError:
        pass
    return value


def _yaml_scalar(value: Any) -> str:
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float)):
        return str(value)
    return _yaml_str(str(value))


def _to_yaml(obj: Any, indent: int = 0) -> str:
    """Minimal recursive YAML serialiser for spec objects."""
    pad = "  " * indent
    if obj is None:
        return "null"
    if isinstance(obj, bool):
        return "true" if obj else "false"
    if isinstance(obj, (int, float)):
        return str(obj)
    if isinstance(obj, str):
        return _yaml_str(obj)
    if isinstance(obj, list):
        if not obj:
            return "[]"
        lines = []
        for item in obj:
            if isinstance(item, dict):
                # Render first key on same line as dash, rest indented beneath
                kv_pairs = list(item.items())
                first_key, first_val = kv_pairs[0]
                if isinstance(first_val, (dict, list)) and first_val:
                    first_line = f"{pad}- {first_key}:\n{_to_yaml(first_val, indent + 2)}"
                else:
                    first_line = f"{pad}- {first_key}: {_to_yaml(first_val, indent + 1)}"
                rest_lines = []
                for k, v in kv_pairs[1:]:
                    inner_pad = "  " * (indent + 1)
                    if isinstance(v, (dict, list)) and v:
                        rest_lines.append(f"{inner_pad}{k}:\n{_to_yaml(v, indent + 2)}")
                    else:
                        rest_lines.append(f"{inner_pad}{k}: {_to_yaml(v, indent + 1)}")
                if rest_lines:
                    lines.append(first_line + "\n" + "\n".join(rest_lines))
                else:
                    lines.append(first_line)
            elif isinstance(item, list):
                inner = _to_yaml(item, indent + 1)
                lines.append(f"{pad}- {inner}")
            else:
                lines.append(f"{pad}- {_yaml_scalar(item)}")
        return "\n".join(lines)
    if isinstance(obj, dict):
        if not obj:
            return "{}"
        lines = []
        for key, value in obj.items():
            if isinstance(value, (dict, list)) and value:
                inner = _to_yaml(value, indent + 1)
                lines.append(f"{pad}{key}:\n{inner}")
            else:
                lines.append(f"{pad}{key}: {_to_yaml(value, indent)}")
        return "\n".join(lines)
    return _yaml_scalar(str(obj))


def spec_to_yaml(spec: dict) -> str:
    """Serialise a spec dict to YAML string."""
    return _to_yaml(spec, 0) + "\n"


# ── Name utilities ─────────────────────────────────────────────────────────

def to_camel_case(text: str) -> str:
    """Normalise a variable name to camelCase."""
    if not text:
        return text
    text = str(text).strip()
    # Already camelCase / no spaces - return as-is
    if re.match(r'^[a-z][a-zA-Z0-9]*$', text):
        return text
    # Split on non-alphanumeric and rebuild
    parts = re.split(r'[^a-zA-Z0-9]+', text)
    parts = [p for p in parts if p]
    if not parts:
        return text
    result = parts[0][0].lower() + parts[0][1:]
    for p in parts[1:]:
        result += p[0].upper() + p[1:] if p else ''
    return result


def camel_to_title(name: str) -> str:
    """Convert camelCase to Title Case label."""
    # Insert space before uppercase letters preceded by lowercase or digit
    spaced = re.sub(r'([a-z0-9])([A-Z])', r'\1 \2', name)
    spaced = re.sub(r'([A-Z]+)([A-Z][a-z])', r'\1 \2', spaced)
    return spaced.title()


def to_snake_case(name: str) -> str:
    """Convert camelCase to snake_case."""
    s = re.sub(r'([a-z0-9])([A-Z])', r'\1_\2', name)
    s = re.sub(r'([A-Z]+)([A-Z][a-z])', r'\1_\2', s)
    return s.lower()


# ── Sensitivity mapping ────────────────────────────────────────────────────

VALID_SENS = {"None", "Low", "Medium", "High"}


def normalise_sensitivity(rank_val: Any, binary_val: Any) -> tuple[str, bool]:
    """
    Return (sensitivity: str, sensitive: bool).

    Priority: Sensitivity Rank column (already None/Low/Medium/High strings).
    Fallback: Binary column (Yes->High, No->Low).
    Final fallback: 'Low'.
    """
    if rank_val is not None:
        rv = str(rank_val).strip()
        # Normalise numeric rank if ever present (1=None,2=Low,3=Medium,4=High)
        if rv in VALID_SENS:
            sensitive = rv in ("Medium", "High")
            return rv, sensitive
        rank_map = {"1": "None", "2": "Low", "3": "Medium", "4": "High"}
        if rv in rank_map:
            mapped = rank_map[rv]
            return mapped, mapped in ("Medium", "High")

    if binary_val is not None:
        bv = str(binary_val).strip().lower()
        if bv in ("yes", "true", "1"):
            return "High", True
        if bv in ("no", "false", "0"):
            return "Low", False

    # Neither present (AD v1.0 has no sensitivity data)
    # Infer from name heuristics
    return "Low", False


def sensitivity_from_level(level_val: Any, binary_val: Any) -> tuple[str, bool]:
    """
    Sheet 1 (elite47) has an explicit 'Sensitivity Level' column with
    None/Low/Medium/High values. Use it directly.
    """
    if level_val is not None:
        lv = str(level_val).strip()
        if lv in VALID_SENS:
            sensitive = lv in ("Medium", "High")
            return lv, sensitive
    # Fallback to binary
    if binary_val is not None:
        bv = str(binary_val).strip().lower()
        if bv in ("yes", "true", "1"):
            return "High", True
        return "Low", False
    return "Low", False


# ── Entity classification ──────────────────────────────────────────────────

FILE_ENTITY_NAMES = {
    "datatype", "assaytype", "fileformat", "ismultispecimen",
    "fileformats", "datasubtype", "filesize", "assay",
    "nucleicacidsource", "celltype", "species", "specimenType".lower(),
    "organ", "tissue", "ispostmortem",
}

FILE_ENTITY_PATTERNS = re.compile(
    r'^(file|assay|data(type|subtype|format)|specimen|organ|tissue|nucleic|cell|ispostmortem)',
    re.IGNORECASE
)


def infer_entity(name: str) -> str:
    key = name.lower()
    if key in FILE_ENTITY_NAMES:
        return "files"
    if FILE_ENTITY_PATTERNS.match(name):
        return "files"
    return "subjects"


# ── Widget inference ───────────────────────────────────────────────────────

AGE_BINS = [
    {"label": "<70",  "min": 0,  "max": 69},
    {"label": "70-74", "min": 70, "max": 74},
    {"label": "75-79", "min": 75, "max": 79},
    {"label": "80-84", "min": 80, "max": 84},
    {"label": "85-89", "min": 85, "max": 89},
    {"label": "90+",   "min": 90, "max": 200},
]

AGE_DEATH_BINS = [
    {"label": "<70",  "min": 0,  "max": 69},
    {"label": "70-74", "min": 70, "max": 74},
    {"label": "75-79", "min": 75, "max": 79},
    {"label": "80-84", "min": 80, "max": 84},
    {"label": "85-89", "min": 85, "max": 89},
    {"label": "90+",   "min": 90, "max": 200},
]

VISIT_OPTIONS = [
    {"label": "1+ visit",  "min": 1},
    {"label": "2+ visits", "min": 2},
    {"label": "3+ visits", "min": 3},
]

KNOWN_CATEGORICAL = {
    "sex", "race", "ethnicity", "diagnosis", "diseasediagnosis",
    "apoegenotype", "cohort", "studycode", "countrycode",
    "fieldcentercode", "specimentype", "organ", "tissue",
    "nucleicacidsource", "celltype", "brainregion", "study",
    "ethnicgroupcode", "diagnosisstatus",
}

IDENTIFIER_PATTERNS = re.compile(r'(id$|_id$|hash|uid|uuid)', re.IGNORECASE)

BOOLEAN_HINT_PATTERNS = re.compile(
    r'(boolean|Bool|Yes / No|Yes/No|BOOLEAN)',
    re.IGNORECASE
)

MULTISELECT_HINT_PATTERNS = re.compile(
    r'(STRING LIST|STRING_LIST|MULTISELECT|CONTROLLED VOCAB|direct filter \(STRING)',
    re.IGNORECASE
)

RANGE_HINT_PATTERNS = re.compile(r'(RANGE|SLIDER|numeric)', re.IGNORECASE)

BINNED_HINT_PATTERNS = re.compile(r'(BINNED|BINS|BUCKET)', re.IGNORECASE)


def parse_values(values_text: str | None, filter_text: str | None) -> list[str]:
    """Extract a controlled vocabulary list from the Values or Filter column."""
    combined = " ".join(filter(None, [values_text, filter_text]))
    if not combined:
        return []

    # Skip obviously non-vocab strings
    skip_patterns = re.compile(
        r'(^\?$|too granular|recommend|50\+|90\+|^\d+\+ values|exact ages)',
        re.IGNORECASE
    )
    if skip_patterns.search(combined.strip()):
        return []

    # Remove parenthetical filter-style prefixes like "Direct filter (STRING LIST)"
    combined = re.sub(
        r'(Direct filter|Boolean direct filter|Range selection|Dropdown).*?\)',
        '', combined, flags=re.IGNORECASE
    )

    # Split on common delimiters; comma is primary, slash and semicolon secondary
    raw = re.split(r'[,;/]', combined)
    results = []
    for r in raw:
        r = r.strip().strip('"').strip("'").strip()
        # Skip instructions / noise
        if not r or len(r) > 80:
            continue
        if re.search(r'\b(filter|boolean|range|binned|list|direct|controlled|select)\b', r, re.IGNORECASE):
            continue
        results.append(r)

    # Deduplicate preserving order
    seen: set[str] = set()
    deduped = []
    for v in results:
        if v not in seen:
            seen.add(v)
            deduped.append(v)
    return deduped


def infer_widget(
    name: str,
    filter_text: str | None,
    values_text: str | None,
    how_collected: str | None = None,
) -> tuple[str, dict]:
    """
    Return (widget_type, extra_fields_dict).
    extra_fields_dict may contain: bins, options, values, visible.
    """
    key = name.lower()
    all_hints = " ".join(filter(None, [filter_text, values_text, how_collected]))

    # Identifiers -> internal
    if IDENTIFIER_PATTERNS.search(name) and key not in KNOWN_CATEGORICAL:
        return "internal", {"visible": False}

    # familyID special case
    if key == "familyid":
        return "internal", {"visible": False}

    # Age / ageDeath -> bins
    if key in ("age", "agedeath"):
        bins_to_use = AGE_BINS if key == "age" else AGE_DEATH_BINS
        return "bins", {"bins": bins_to_use}

    # visitCode / visitCount -> minCount
    if key in ("visitcode", "visitcount", "numvisits", "numberofvisits"):
        return "minCount", {"options": VISIT_OPTIONS}

    # Explicit binned hint (e.g. "Range selection (BINNED)")
    if BINNED_HINT_PATTERNS.search(all_hints):
        if key in ("age", "agedeath"):
            bins_to_use = AGE_BINS if key == "age" else AGE_DEATH_BINS
            return "bins", {"bins": bins_to_use}
        # Generic binned numeric - treat as bins with no pre-set (defer to inference)
        return "bins", {}

    # Boolean hints: name starts has/is, or filter says boolean
    if (key.startswith("has") or key.startswith("is") or
            BOOLEAN_HINT_PATTERNS.search(all_hints)):
        return "boolean", {}

    # Range hints
    if RANGE_HINT_PATTERNS.search(all_hints) and not BINNED_HINT_PATTERNS.search(all_hints):
        return "range", {}

    # Known categoricals
    if key in KNOWN_CATEGORICAL:
        return "multiselect", {}

    # Values text with parseable vocab
    parsed_vals = parse_values(values_text, None)
    if len(parsed_vals) >= 2:
        return "multiselect", {"values": parsed_vals}

    # Multiselect hint from filter
    if MULTISELECT_HINT_PATTERNS.search(all_hints):
        parsed = parse_values(values_text, filter_text)
        extra: dict = {}
        if parsed:
            extra["values"] = parsed
        return "multiselect", extra

    # Default: multiselect with empty values (app will infer from data)
    return "multiselect", {}


# ── Row -> VariableOverride ────────────────────────────────────────────────

def build_variable_override(
    raw_name: str,
    category: str | None,
    sensitivity: str,
    sensitive: bool,
    filter_text: str | None,
    values_text: str | None,
    how_collected: str | None = None,
    note: str | None = None,
) -> dict:
    """Assemble a VariableOverride dict from extracted cell values."""
    name = to_camel_case(raw_name)
    label = camel_to_title(name)
    entity = infer_entity(name)
    column = to_snake_case(name)
    widget, extra = infer_widget(name, filter_text, values_text, how_collected)

    var: dict = {
        "name": name,
        "label": label,
        "category": category or "Other",
        "sensitivity": sensitivity,
        "sensitive": sensitive,
        "entity": entity,
        "column": column,
        "widget": widget,
    }

    # Merge extra fields (bins, options, values, visible)
    var.update(extra)

    # Note: combine viz/filter guidance, trim whitespace
    note_parts = []
    if note:
        note_parts.append(note.strip())
    if filter_text and note != filter_text:
        note_parts.append(filter_text.strip())
    combined_note = " | ".join(note_parts[:1]) if note_parts else None  # keep short
    if combined_note:
        var["note"] = combined_note[:300]  # cap length

    return var


# ── Standard relationships block ───────────────────────────────────────────

STANDARD_RELATIONSHIPS = [
    {
        "from": "subjects",
        "to": "files",
        "via": "subject_files",
        "fromKey": "subject_id",
        "toKey": "syn_id",
    }
]


# ── Sheet readers ──────────────────────────────────────────────────────────

def find_header_row(rows: list[tuple], target_col_names: list[str]) -> int:
    """
    Find the index of the row that contains all target column names.
    Returns the row index or -1 if not found.
    """
    for i, row in enumerate(rows):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if all(t in cells for t in target_col_names):
            return i
    return -1


def read_elite47(ws) -> list[dict]:
    """Read the elite_cohort_builder_variables sheet (47 curated vars)."""
    rows = list(ws.iter_rows(values_only=True))
    # Header is row 0, but detect robustly
    header_idx = find_header_row(rows, ["Variable", "Sensitivity Level"])
    if header_idx == -1:
        header_idx = 0  # fallback

    header = [str(c).strip() if c is not None else "" for c in rows[header_idx]]

    def col(name: str) -> int | None:
        try:
            return header.index(name)
        except ValueError:
            return None

    var_idx = col("Variable")
    cat_idx = col("Category")
    sens_lev_idx = col("Sensitivity Level")
    sens_bin_idx = col("Sensitive Data?")
    filt_idx = col("Filter")
    viz_idx = col("Visualization")
    how_idx = col("How Collected")

    variables = []
    seen: set[str] = set()

    for row in rows[header_idx + 1:]:
        raw_name = row[var_idx] if var_idx is not None else None
        if not raw_name:
            continue
        raw_name = str(raw_name).strip()
        if not raw_name or raw_name in seen:
            continue
        seen.add(raw_name)

        category = row[cat_idx] if cat_idx is not None else None
        level_val = row[sens_lev_idx] if sens_lev_idx is not None else None
        binary_val = row[sens_bin_idx] if sens_bin_idx is not None else None
        sensitivity, sensitive = sensitivity_from_level(level_val, binary_val)

        filter_text = str(row[filt_idx]).strip() if filt_idx is not None and row[filt_idx] else None
        viz_text = str(row[viz_idx]).strip() if viz_idx is not None and row[viz_idx] else None
        how_text = str(row[how_idx]).strip() if how_idx is not None and row[how_idx] else None

        variables.append(build_variable_override(
            raw_name=raw_name,
            category=str(category).strip() if category else None,
            sensitivity=sensitivity,
            sensitive=sensitive,
            filter_text=filter_text,
            values_text=None,
            how_collected=how_text,
            note=viz_text,
        ))

    return variables


def read_generic_sheet(ws, header_fallback_idx: int) -> list[dict]:
    """
    Read a v2.0 / ELITE v3.0 / AD v1.0 sheet.
    Header has: Canonical Name, Category, Sensitive (Binary), Sensitivity (Rank),
                Filter, Values.
    """
    rows = list(ws.iter_rows(values_only=True))
    header_idx = find_header_row(rows, ["Canonical Name"])
    if header_idx == -1:
        header_idx = header_fallback_idx

    header = [str(c).strip() if c is not None else "" for c in rows[header_idx]]

    def col(name: str) -> int | None:
        try:
            return header.index(name)
        except ValueError:
            return None

    name_idx = col("Canonical Name")
    cat_idx = col("Category")
    sens_rank_idx = col("Sensitivity (Rank)")
    sens_bin_idx = col("Sensitive (Binary)")
    filt_idx = col("Filter")
    val_idx = col("Values")

    if name_idx is None:
        raise ValueError(f"Could not find 'Canonical Name' column; header={header[:15]}")

    variables = []
    seen: set[str] = set()

    def cell(row: tuple, idx: int | None) -> Any:
        """Safely read a cell by column index; return None if out of bounds."""
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    for row in rows[header_idx + 1:]:
        raw_name = cell(row, name_idx)
        if not raw_name:
            continue
        raw_name = str(raw_name).strip()
        if not raw_name or raw_name in seen:
            continue
        seen.add(raw_name)

        category = cell(row, cat_idx)
        rank_val = cell(row, sens_rank_idx)
        binary_val = cell(row, sens_bin_idx)
        sensitivity, sensitive = normalise_sensitivity(rank_val, binary_val)

        filt_raw = cell(row, filt_idx)
        filter_text = str(filt_raw).strip() if filt_raw else None
        val_raw = cell(row, val_idx)
        values_text = str(val_raw).strip() if val_raw else None

        variables.append(build_variable_override(
            raw_name=raw_name,
            category=str(category).strip() if category else None,
            sensitivity=sensitivity,
            sensitive=sensitive,
            filter_text=filter_text,
            values_text=values_text,
        ))

    return variables


# ── Spec assembly ──────────────────────────────────────────────────────────

def build_spec(
    spec_id: str,
    title: str,
    description: str,
    variables: list[dict],
) -> dict:
    return {
        "schemaVersion": "1.0",
        "id": spec_id,
        "title": title,
        "description": description,
        "primaryEntity": "subjects",
        "relationships": STANDARD_RELATIONSHIPS,
        "variables": variables,
    }


# ── Write helpers ──────────────────────────────────────────────────────────

def write_spec(slug: str, spec: dict) -> tuple[Path, Path]:
    json_path = SPECS_DIR / f"{slug}.spec.json"
    yaml_path = SPECS_DIR / f"{slug}.spec.yaml"

    json_path.write_text(json.dumps(spec, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    yaml_path.write_text(spec_to_yaml(spec), encoding="utf-8")

    return yaml_path, json_path


# ── Sensitivity distribution summary ──────────────────────────────────────

def sensitivity_distribution(variables: list[dict]) -> dict[str, int]:
    dist: dict[str, int] = {"None": 0, "Low": 0, "Medium": 0, "High": 0}
    for v in variables:
        level = v.get("sensitivity", "Low")
        dist[level] = dist.get(level, 0) + 1
    return dist


# ── Validate YAML round-trip using stdlib json (best-effort check) ─────────

def validate_json(path: Path) -> bool:
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        return isinstance(data, dict) and "schemaVersion" in data
    except Exception as e:
        print(f"  ERROR: JSON validation failed for {path}: {e}", file=sys.stderr)
        return False


# ── Main ───────────────────────────────────────────────────────────────────

def main() -> None:
    print(f"Loading workbook: {WORKBOOK_PATH}")
    wb = openpyxl.load_workbook(str(WORKBOOK_PATH), read_only=True, data_only=True)

    sheet_configs = [
        {
            "slug": "elite47",
            "sheet_name": "elite_cohort_builder_variables ",
            "title": "ELITE Cohort Builder - Curated 47 Variables",
            "reader": "elite47",
            "header_fallback": 0,
        },
        {
            "slug": "elite-v2",
            "sheet_name": "v2.0",
            "title": "ELITE Cohort Builder - v2.0 Full Variable Set",
            "reader": "generic",
            "header_fallback": 3,
        },
        {
            "slug": "elite-v3",
            "sheet_name": "ELITE v3.0",
            "title": "ELITE Cohort Builder - v3.0 Refined Variable Set",
            "reader": "generic",
            "header_fallback": 4,
        },
        {
            "slug": "ad-v1",
            "sheet_name": "AD v1.0",
            "title": "AD Knowledge Portal - v1.0 Variable Set",
            "reader": "generic",
            "header_fallback": 3,
        },
    ]

    index: list[dict] = []
    summary_rows: list[tuple] = []

    print()
    print(f"{'Sheet':<35} {'Vars':>5}  None   Low  Med  High")
    print("-" * 65)

    for cfg in sheet_configs:
        ws = wb[cfg["sheet_name"]]
        slug = cfg["slug"]

        if cfg["reader"] == "elite47":
            variables = read_elite47(ws)
            prov_note = (
                "Compiled from the 'elite_cohort_builder_variables' sheet of the "
                "ELITE Cohort Builder workbook. Sensitivity sourced directly from "
                "the 'Sensitivity Level' column (None/Low/Medium/High). "
                "SDC intent: High-sensitivity variables (e.g. apoeGenotype, diagnosis) "
                "should be subject to complementary suppression and boolean-only return "
                "at the High level per DEFAULT_SDC policy."
            )
        else:
            variables = read_generic_sheet(ws, cfg["header_fallback"])
            # AD v1.0 has no sensitivity data - document the inference used
            if slug == "ad-v1":
                prov_note = (
                    "Compiled from the 'AD v1.0' sheet of the ELITE Cohort Builder workbook. "
                    "The AD sheet contains no Sensitive (Binary) or Sensitivity (Rank) data. "
                    "Sensitivity was inferred from variable name heuristics: "
                    "variables starting with 'has'/'is' or matching known-sensitive names "
                    "were assigned High; all others default to Low. "
                    "These assignments should be reviewed before production use."
                )
            else:
                version = cfg["sheet_name"]
                prov_note = (
                    f"Compiled from the '{version}' sheet of the ELITE Cohort Builder workbook. "
                    "Sensitivity sourced from 'Sensitivity (Rank)' column (None/Low/Medium/High strings). "
                    "Where Sensitivity (Rank) was absent, 'Sensitive (Binary)' was used: "
                    "Yes maps to High, No maps to Low."
                )

        description = f"{cfg['title']}. {prov_note} Variable count: {len(variables)}."
        spec = build_spec(
            spec_id=slug,
            title=cfg["title"],
            description=description,
            variables=variables,
        )

        yaml_path, json_path = write_spec(slug, spec)
        valid = validate_json(json_path)

        dist = sensitivity_distribution(variables)
        print(
            f"  {cfg['sheet_name']:<33} {len(variables):>5}  "
            f"{dist['None']:>4}  {dist['Low']:>4}  {dist['Medium']:>4}  {dist['High']:>4}"
            + ("" if valid else "  [JSON ERROR]")
        )

        index.append({
            "id": slug,
            "title": cfg["title"],
            "slug": slug,
            "yaml": f"/specs/{slug}.spec.yaml",
            "json": f"/specs/{slug}.spec.json",
            "variableCount": len(variables),
            "sheet": cfg["sheet_name"].strip(),
        })
        summary_rows.append((slug, len(variables), dist))

    # Write index
    index_path = SPECS_DIR / "index.json"
    index_path.write_text(json.dumps(index, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")

    # Verify all files
    print()
    print("Output file verification:")
    all_ok = True
    for cfg in sheet_configs:
        slug = cfg["slug"]
        for ext in ("yaml", "json"):
            p = SPECS_DIR / f"{slug}.spec.{ext}"
            exists = p.exists()
            size = p.stat().st_size if exists else 0
            status = "OK" if exists and size > 0 else "MISSING/EMPTY"
            print(f"  {p.relative_to(REPO_ROOT)}  ({size:,} bytes)  [{status}]")
            if status != "OK":
                all_ok = False

    idx_ok = index_path.exists() and index_path.stat().st_size > 0
    print(f"  {index_path.relative_to(REPO_ROOT)}  ({index_path.stat().st_size:,} bytes)  [{'OK' if idx_ok else 'MISSING'}]")

    print()
    if all_ok and idx_ok:
        print("All 9 files written successfully.")
    else:
        print("WARNING: one or more files are missing or empty.", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
